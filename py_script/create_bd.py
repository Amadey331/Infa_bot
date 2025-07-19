

import openpyxl
import sqlite3

import os




async def create_bdL():
    path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'bd_les\\Group_les_bd_up.db')
    os.remove(path)
    path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'bd_les\\Group_les_bd_low.db')
    os.remove(path)

    bdGroupLes_up= sqlite3.connect(r"py_script/bd_les/Group_les_bd_up.db")
    bdGroupLes_low= sqlite3.connect(r"py_script/bd_les/Group_les_bd_low.db")
    cur_up= bdGroupLes_up.cursor()
    cur_low= bdGroupLes_low.cursor()
    days = ("Понедельник","Вторник","Среда","Четверг","Пятница","Суббота","Воскресенье")
    file = openpyxl.open(r"py_script/les.xlsx" , read_only = True)
    sheet = file.active
    for i in range(sheet.max_column):
        

        if  sheet[1][i].value!=None  and "Группа" in sheet[1][i].value:
            list_forUpgroup = []
            list_forLowgroup = []

            string_group_ful = sheet[1][i].value
            string_group_end = string_group_ful.find("(") 
            result=string_group_ful[:string_group_end-1].replace(" ","_")
            
            bdGroupLes_up.execute("CREATE TABLE IF NOT EXISTS  {}(day,time,name,location,room_number,type)".format( result))
            bdGroupLes_up.commit()
            bdGroupLes_low.execute("CREATE TABLE IF NOT EXISTS {}(day,time,name,location,room_number,type)".format( result))
            bdGroupLes_low.commit()
            daysName = ""
            
            for row in sheet.iter_rows(min_row=3,max_row=sheet.max_row,min_col=i-2,max_col=i+4):
                
                add_list = []
                for cell in row:
                    if cell.value in days:
                        daysName = cell.value
                        add_list.append(daysName)
                    else:
                        add_list.append(str(cell.value))
                        add_list[0] = daysName
                
                
                if add_list [2].lower()=="в":
                    add_list.remove("в")
                    list_forUpgroup.append(add_list)
                elif add_list [2].lower()=="н":
                    add_list.remove("н")
                    list_forLowgroup.append(add_list)
            cur_up.executemany("INSERT INTO {} VALUES(?,?,?,?,?,?)".format(result),(list_forUpgroup))
            bdGroupLes_up.commit()
            cur_low.executemany("INSERT INTO {} VALUES(?,?,?,?,?,?)".format(result),(list_forLowgroup))
            bdGroupLes_low.commit()    








    